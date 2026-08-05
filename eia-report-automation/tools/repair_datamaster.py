"""마스터DB·입력양식·예시데이터 일괄 정비 (v6 → v7).

v6 에서 확인한 결함을 고치고, 세 파일을 `species_id` 로 연동한다.

정비 항목
  1. species_id 연동 — 입력양식·예시데이터에 조인 키를 넣는다. v6 은 행
     위치로만 조인할 수 있어 사용자가 정렬·삽입하면 깨졌다.
  2. 학명 표기 통일 — 유니코드 수학 이탤릭을 ASCII 로 정규화한다. 이탤릭은
     데이터가 아니라 표현이므로 셀 서식으로 처리한다.
  3. 분류체계 결측 보정 — 속(genus)이 같은 다른 행에서 문·강·목·과를
     채운다. 추론이 불가능하면 비워 둔다.
  4. 중복 행 병합 — 학명·국명이 같은 행을 하나로 합친다. 값이 충돌하면
     첫 값을 쓰고 충돌 내역을 기록한다.
  5. abb 재계산 — 병합 후 구성요소로 다시 만든다.

원칙: 없는 정보를 지어내지 않는다. 국명이 없는 종은 전문가 확인 대상으로
남기고, 병합 시 버린 값은 반드시 기록한다.

실행:
    python -m tools.repair_datamaster
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "datamaster"

SRC_MASTER = DATA / "EIA_표준종목록_마스터DB_통합본_v6.xlsx"
SRC_INPUT = DATA / "EIA_표준종목록_입력용_데이터시트_v6.xlsx"
SRC_SAMPLE = DATA / "EIA_가상데이터.xlsx"

OUT_MASTER = DATA / "EIA_표준종목록_마스터DB_통합본_v7.xlsx"
OUT_INPUT = DATA / "EIA_표준종목록_입력용_데이터시트_v7.xlsx"
OUT_SAMPLE = DATA / "EIA_가상데이터_v7.xlsx"

TAXA = ["포유류", "조류", "양서류", "파충류", "어류",
        "육상곤충류", "저서성대형무척추동물", "관속식물"]

SURVEY_COLS = ["문헌1", "문헌2", "현지조사1", "현지조사2"]
INPUT_COLS = ["species_id", "family_kr", "scientific_name", "korean_name", "abb", "abb2"]

# 분류체계 컬럼. 속이 같으면 같은 값을 갖는다.
RANK_COLS = ["phylum", "phylum_kr", "class", "class_kr", "order", "order_kr",
             "family", "family_kr"]

MISSING_KOREAN = "[국명없음]"
NULL = "-"

# v6 에는 국명 결측 표기가 세 가지로 흩어져 있다. 하나로 통일한다.
MISSING_KOREAN_VARIANTS = {"[국명없음]", "국명없음", "국명미정", "국명 없음", "국명 미정"}

# 학명에 섞인 EN DASH 는 하이픈 오타다(Word 자동 교정 흔적).
# 잡종 기호 '×'(U+00D7)는 식물 명명법상 정상이므로 건드리지 않는다.
DASH_VARIANTS = {"–": "-", "—": "-", "−": "-"}


def is_text_column(series: pd.Series) -> bool:
    """문자열 컬럼인가.

    pandas 2.x 는 문자열 컬럼의 dtype 을 object 가 아니라 str 로 추론한다.
    `dtype == object` 로 판별하면 정규화가 통째로 건너뛰어진다.
    """
    return pd.api.types.is_string_dtype(series) or series.dtype == object


# 등급·코드 컬럼. 로마숫자(Ⅰ~Ⅴ)를 값으로 쓰므로 NFKC 를 적용하면 안 된다.
# NFKC 는 호환 문자를 분해하기 때문에 'Ⅱ'(U+2161) 가 'II'(ASCII 2자) 로 바뀐다.
CODE_COLUMNS = {
    "species_id", "taxon_rank", "nibr_ktsn", "source_flag",
    "멸종위기야생생물", "천연기념물", "생태계교란생물", "고유종", "외래종",
    "식물구계학적특정종", "희귀식물등급", "특산식물", "귀화식물",
    "abb", "abb2", "raunkiaer_form", "migratory_type",
    "saprobic_index_Qi", "dual_habitat_ref_id", "merged_from",
}


def norm(value: object, keep_compat: bool = False) -> str:
    """공백 정리 + 유니코드 정규화.

    기본은 NFKC 로, 학명의 유니코드 수학 이탤릭('𝐸𝑟𝑖𝑛𝑎𝑐𝑒𝑢𝑠')을 ASCII 로
    되돌린다. `keep_compat=True` 면 NFC 를 써서 호환 문자를 분해하지
    않는다. 등급 컬럼의 로마숫자를 보존하기 위한 것이다.
    """
    if value is None:
        return ""
    form = "NFC" if keep_compat else "NFKC"
    return " ".join(unicodedata.normalize(form, str(value)).split())


def norm_column(value: object, column: str) -> str:
    """컬럼 성격에 맞는 정규화를 고른다."""
    return norm(value, keep_compat=column in CODE_COLUMNS)


def is_null(value: object) -> bool:
    return norm(value, keep_compat=True) in ("", NULL, "nan", "None", "NaN")


def clean(value: object, column: str | None = None) -> str:
    """값이 없으면 '-' 로 통일한다."""
    if is_null(value):
        return NULL
    return norm(value, keep_compat=column in CODE_COLUMNS) if column else norm(value)


@dataclass
class RepairLog:
    """정비 내역. 무엇을 왜 바꿨는지 남긴다."""

    rows: list[dict] = field(default_factory=list)

    def add(self, taxon: str, kind: str, target: str, detail: str) -> None:
        self.rows.append({"분류군": taxon, "구분": kind, "대상": target, "내용": detail})

    def frame(self) -> pd.DataFrame:
        if not self.rows:
            return pd.DataFrame(columns=["분류군", "구분", "대상", "내용"])
        return pd.DataFrame(self.rows)


# ── 1. 학명·표기 정규화 ────────────────────────────────────────────────

GENUS_LOWER = re.compile(r"^[a-z]")


def normalize_names(df: pd.DataFrame, taxon: str, log: RepairLog) -> pd.DataFrame:
    for col in df.columns:
        if is_text_column(df[col]):
            df[col] = df[col].map(lambda v, c=col: norm_column(v, c) if v is not None else v)

    unified: dict[str, int] = {}
    for i, r in df.iterrows():
        sid = str(r["species_id"])

        sn = str(r["scientific_name"])
        fixed = sn
        for bad, good in DASH_VARIANTS.items():
            fixed = fixed.replace(bad, good)
        # 속명 첫 글자는 대문자다. 소문자로 시작하면 표기 오류로 본다.
        if fixed and GENUS_LOWER.match(fixed):
            fixed = fixed[0].upper() + fixed[1:]
        if fixed != sn:
            df.at[i, "scientific_name"] = fixed
            log.add(taxon, "학명 표기", sid, f"'{sn}' → '{fixed}'")

        # 국명 결측 표기를 하나로 모은다. 이름을 지어내지는 않는다.
        kn = norm(r["korean_name"])
        if kn in MISSING_KOREAN_VARIANTS and kn != MISSING_KOREAN:
            df.at[i, "korean_name"] = MISSING_KOREAN
            unified[kn] = unified.get(kn, 0) + 1

    # 기계적·일률적 변환이라 건별이 아니라 표기별로 집계해 남긴다
    for old, count in unified.items():
        log.add(taxon, "국명 표기 통일", old, f"'{old}' → '{MISSING_KOREAN}' {count}건")
    return df


# ── 2. 분류체계 결측 보정 ──────────────────────────────────────────────

def fill_taxonomy(df: pd.DataFrame, taxon: str, log: RepairLog) -> pd.DataFrame:
    """속이 같은 다른 행에서 문·강·목·과를 채운다."""
    if "genus" not in df.columns:
        return df

    # 속명이 비어 있으면 학명 첫 단어에서 얻는다
    for i, r in df.iterrows():
        if is_null(r.get("genus")) and str(r["scientific_name"]).strip():
            df.at[i, "genus"] = str(r["scientific_name"]).split()[0]

    ref: dict[str, dict[str, str]] = {}
    for _, r in df.iterrows():
        g = norm(r.get("genus"))
        if not g or is_null(r.get("family_kr")):
            continue
        ref.setdefault(g, {c: clean(r.get(c)) for c in RANK_COLS if c in df.columns})

    filled = 0
    for i, r in df.iterrows():
        if not is_null(r.get("family_kr")):
            continue
        src = ref.get(norm(r.get("genus")))
        if not src:
            log.add(taxon, "분류체계 미보정", str(r["species_id"]),
                    f"{r['scientific_name']} — 같은 속의 참조 행이 없어 비워 둠")
            continue
        for c, v in src.items():
            if is_null(df.at[i, c]):
                df.at[i, c] = v
        filled += 1
        log.add(taxon, "분류체계 보정", str(r["species_id"]),
                f"{r['scientific_name']} — 속 '{r['genus']}' 기준으로 {src.get('family_kr')} 채움")
    if filled:
        log.add(taxon, "요약", "분류체계 보정", f"{filled}건")
    return df


# ── 3. 계급(taxon_rank) 보정 ───────────────────────────────────────────

RANK_CONNECTORS = ("var.", "f.", "subsp.", "ssp.", "×", "x ")


def fix_rank(df: pd.DataFrame, taxon: str, log: RepairLog) -> pd.DataFrame:
    """연결어 없는 3단어 학명은 아종이다. DB 자체 관례를 따른다."""
    if "taxon_rank" not in df.columns:
        return df
    n = 0
    for i, r in df.iterrows():
        sn = str(r["scientific_name"])
        if norm(r.get("taxon_rank"), keep_compat=True) != "종":
            continue
        if any(c in sn for c in RANK_CONNECTORS):
            continue
        if len(sn.split()) == 3:
            df.at[i, "taxon_rank"] = "아종"
            n += 1
            log.add(taxon, "계급 보정", str(r["species_id"]), f"{sn} — 종 → 아종(삼명법)")
    if n:
        log.add(taxon, "요약", "계급 보정", f"{n}건")
    return df


# ── 4. 중복 병합 ───────────────────────────────────────────────────────

def merge_duplicates(df: pd.DataFrame, taxon: str, log: RepairLog
                     ) -> tuple[pd.DataFrame, dict[str, str]]:
    """학명·국명이 같은 행을 하나로 합친다.

    반환값의 두 번째 항목은 {버려진 species_id: 살아남은 species_id} 로,
    예시데이터의 출현 기록을 옮길 때 쓴다.
    """
    key = ["scientific_name", "korean_name"]
    dup_keys = df[df.duplicated(key, keep=False)].groupby(key).groups
    if not dup_keys:
        return df, {}

    attrs = [c for c in df.columns if c not in ("species_id", *key)]
    remap: dict[str, str] = {}
    drop_idx: list[int] = []

    for _, idx in dup_keys.items():
        idx = list(idx)
        keep, rest = idx[0], idx[1:]
        keep_id = str(df.at[keep, "species_id"])
        for c in attrs:
            values, seen = [], set()
            for i in idx:
                v = clean(df.at[i, c], c)
                if v != NULL and v not in seen:
                    seen.add(v)
                    values.append(v)
            if not values:
                continue
            df.at[keep, c] = values[0]
            if len(values) > 1:
                log.add(taxon, "병합 충돌", keep_id,
                        f"{df.at[keep, 'scientific_name']} · {c}: "
                        f"'{values[0]}' 채택, 버림 {values[1:]}")
        merged_ids = [str(df.at[i, "species_id"]) for i in rest]
        for mid in merged_ids:
            remap[mid] = keep_id
        df.at[keep, "merged_from"] = "/".join(merged_ids)
        drop_idx.extend(rest)
        log.add(taxon, "중복 병합", keep_id,
                f"{df.at[keep, 'scientific_name']} · {df.at[keep, 'korean_name']} "
                f"— {merged_ids} 흡수")

    df = df.drop(index=drop_idx).reset_index(drop=True)
    log.add(taxon, "요약", "중복 병합", f"{len(drop_idx)}행 제거, {len(dup_keys)}그룹으로 통합")
    return df, remap


# ── 4.5 수서곤충 교차 참조 복구 ────────────────────────────────────────

# 육상곤충류와 저서성대형무척추동물은 같은 종을 유충(수서)·성충(육상)으로
# 나눠 싣는다. dual_habitat_ref_id 가 그 짝을 가리키는데, v6 은 ID 형식이
# 어긋나 한 건도 해결되지 않는다.
#   저서 시트의 참조: 'IN08312'  → 곤충 ID 는 'IN000001' (6자리)
#   곤충 시트의 참조: 'BI00500'  → 저서 ID 는 'BE00001'  (접두사 BE)
DUAL_HABITAT_PAIR = {"육상곤충류": ("저서성대형무척추동물", "BI", "BE", 5),
                     "저서성대형무척추동물": ("육상곤충류", "IN", "IN", 6)}


def fix_dual_habitat(sheets: dict[str, pd.DataFrame], log: RepairLog) -> None:
    """교차 참조 ID 를 상대 시트의 실제 형식으로 맞춘다."""
    for taxon, (other, src_prefix, dst_prefix, width) in DUAL_HABITAT_PAIR.items():
        df = sheets.get(taxon)
        if df is None or "dual_habitat_ref_id" not in df.columns:
            continue
        valid = set(sheets[other]["species_id"])
        fixed = dangling = 0
        for i, ref in df["dual_habitat_ref_id"].items():
            if is_null(ref):
                continue
            digits = norm(ref, keep_compat=True)[len(src_prefix):]
            if not digits.isdigit():
                continue
            candidate = f"{dst_prefix}{digits.zfill(width)}"
            if candidate in valid:
                if candidate != norm(ref, keep_compat=True):
                    df.at[i, "dual_habitat_ref_id"] = candidate
                    fixed += 1
            else:
                # 상대 시트에 없는 종을 가리킨다. 끊어진 참조를 남겨두면
                # 조인 시 조용히 누락되므로 비운다.
                df.at[i, "dual_habitat_ref_id"] = NULL
                dangling += 1
        if fixed:
            log.add(taxon, "교차 참조 복구", "dual_habitat_ref_id",
                    f"{other} 참조 {fixed}건을 실제 ID 형식으로 보정")
        if dangling:
            log.add(taxon, "확인 필요", "dual_habitat_ref_id",
                    f"{other} 에 대상이 없는 참조 {dangling}건 — 비움")


# ── 5. abb 재계산 ──────────────────────────────────────────────────────

def build_abb(row: pd.Series, is_plant: bool) -> str:
    """법정 지위 통합 약어. v6 설명 시트의 조합 규칙을 따른다."""
    parts: list[str] = []
    if not is_null(row.get("멸종위기야생생물")):
        parts.append(norm(row["멸종위기야생생물"], keep_compat=True))
    if not is_null(row.get("천연기념물")):
        parts.append("천")
    if not is_null(row.get("생태계교란생물")):
        parts.append("교")
    if is_plant:
        if not is_null(row.get("식물구계학적특정종")):
            parts.append(norm(row["식물구계학적특정종"], keep_compat=True))
        if not is_null(row.get("희귀식물등급")):
            parts.append(norm(row["희귀식물등급"], keep_compat=True))
        if not is_null(row.get("특산식물")):
            parts.append("특" if norm(row["특산식물"], keep_compat=True) == "Y" else "특?")
        if not is_null(row.get("귀화식물")):
            parts.append("귀")
    return "/".join(parts) if parts else NULL


def recompute_abb(df: pd.DataFrame, taxon: str, log: RepairLog) -> pd.DataFrame:
    is_plant = taxon == "관속식물"
    changed = 0
    for i, r in df.iterrows():
        new = build_abb(r, is_plant)
        if new != clean(r.get("abb"), "abb"):
            if not is_null(r.get("abb")):
                log.add(taxon, "abb 재계산", str(r["species_id"]),
                        f"'{clean(r.get('abb'), 'abb')}' → '{new}'")
                changed += 1
            df.at[i, "abb"] = new
    if changed:
        log.add(taxon, "요약", "abb 재계산", f"{changed}건 변경")
    return df


# ── 6. 시트 정비 ───────────────────────────────────────────────────────

def repair_sheet(master: pd.DataFrame, taxon: str, log: RepairLog
                 ) -> tuple[pd.DataFrame, dict[str, str]]:
    df = master.copy()
    df["merged_from"] = NULL
    df = normalize_names(df, taxon, log)
    df = fill_taxonomy(df, taxon, log)
    df = fix_rank(df, taxon, log)
    df, remap = merge_duplicates(df, taxon, log)
    df = recompute_abb(df, taxon, log)

    for c in df.columns:
        if is_text_column(df[c]):
            df[c] = df[c].map(lambda v, col=c: clean(v, col))

    missing = int((df["korean_name"] == MISSING_KOREAN).sum())
    if missing:
        log.add(taxon, "확인 필요", "국명 결측",
                f"{missing}건 — 지어내지 않고 그대로 둠. 전문가 확인 필요")
    return df, remap


def merge_survey(sample: pd.DataFrame, remap: dict[str, str],
                 id_by_row: list[str]) -> pd.DataFrame:
    """병합으로 사라진 행의 출현 기록을 살아남은 행으로 옮긴다."""
    sample = sample.copy()
    sample["species_id"] = id_by_row
    if not remap:
        return sample

    by_id = {sid: i for i, sid in enumerate(sample["species_id"])}
    for dropped, kept in remap.items():
        di, ki = by_id.get(dropped), by_id.get(kept)
        if di is None or ki is None:
            continue
        for col in SURVEY_COLS:
            dv, kv = sample.at[di, col], sample.at[ki, col]
            if is_null(dv):
                continue
            if is_null(kv):
                sample.at[ki, col] = dv
            elif str(dv).replace(".0", "").isdigit() and str(kv).replace(".0", "").isdigit():
                # 개체수는 합산한다
                sample.at[ki, col] = int(float(dv)) + int(float(kv))
            elif norm(dv) not in norm(kv):
                sample.at[ki, col] = f"{norm(kv)}/{norm(dv)}"
    return sample[~sample["species_id"].isin(remap)].reset_index(drop=True)


# ── 7. 파일 쓰기 ───────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="E3EFE7")
ID_FILL = PatternFill("solid", fgColor="F2F2F2")


def _style(writer, sheet: str, df: pd.DataFrame) -> None:
    ws = writer.sheets[sheet]
    ws.freeze_panes = "A2"
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {"species_id": 11, "scientific_name": 30, "korean_name": 22,
              "family_kr": 16, "abb": 14, "abb2": 30, "merged_from": 16}
    for i, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 13)

    # 학명은 서식으로 이탤릭을 준다. 문자 자체를 바꾸지 않는다.
    if "scientific_name" in df.columns:
        c = list(df.columns).index("scientific_name") + 1
        for r in range(2, len(df) + 2):
            ws.cell(row=r, column=c).font = Font(italic=True)
    # 조인 키는 수정 대상이 아님을 색으로 알린다
    if "species_id" in df.columns:
        c = list(df.columns).index("species_id") + 1
        for r in range(2, len(df) + 2):
            ws.cell(row=r, column=c).fill = ID_FILL


def _guide(kind: str) -> pd.DataFrame:
    common = [
        ("species_id", "마스터DB와 연결되는 조인 키. 수정·삭제·정렬하지 마십시오."),
        ("", "행을 추가하려면 마스터DB에 종을 먼저 등록해 species_id 를 받으십시오."),
        ("학명 표기", "ASCII 로 통일했습니다. 이탤릭은 셀 서식으로 표시됩니다."),
    ]
    rules = [
        ("입력 규칙", ""),
        ("관속식물·양서류·파충류·육상곤충류", "문헌1·2, 현지조사1·2 모두 출현=1 / 비출현=빈칸"),
        ("조류·어류·저서성대형무척추동물", "문헌1·2는 출현=1. 현지조사1·2는 관찰 개체수(숫자)"),
        ("포유류", "문헌1·2는 출현=1. 현지조사1·2는 조사방법 약어"),
        ("포유류 조사방법", "SI 목견 / TR 족적 / SC 배설물 / AU 청문 / CT 카메라트랩"),
        ("", "복수 검출 시 '/' 로 결합 (예: SI/CT)"),
    ]
    rows = [("EIA 표준종목록 " + kind + " v7", "")] + common
    if kind != "마스터DB":
        rows += [("", "")] + rules
    return pd.DataFrame(rows, columns=["항목", "내용"])


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame],
                   guide: pd.DataFrame | None = None,
                   log_df: pd.DataFrame | None = None) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        if guide is not None:
            guide.to_excel(w, sheet_name="사용법", index=False)
            _style(w, "사용법", guide)
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name, index=False)
            _style(w, name, df)
        if log_df is not None and not log_df.empty:
            log_df.to_excel(w, sheet_name="정비이력", index=False)
            _style(w, "정비이력", log_df)


def main() -> None:
    log = RepairLog()
    masters: dict[str, pd.DataFrame] = {}
    inputs: dict[str, pd.DataFrame] = {}
    samples: dict[str, pd.DataFrame] = {}

    for taxon in TAXA:
        src = pd.read_excel(SRC_MASTER, sheet_name=taxon)
        original_ids = [str(v) for v in src["species_id"]]
        sample_src = pd.read_excel(SRC_SAMPLE, sheet_name=taxon)

        if len(sample_src) != len(src):
            raise ValueError(f"[{taxon}] v6 행 수 불일치: 마스터 {len(src)} / 예시 {len(sample_src)}")

        repaired, remap = repair_sheet(src, taxon, log)
        masters[taxon] = repaired

        # 입력양식: 조인 키 + 표시용 컬럼
        inputs[taxon] = repaired[INPUT_COLS].copy()

        # 예시데이터: 조인 키 + 표시용 컬럼 + 출현 기록
        merged = merge_survey(sample_src, remap, original_ids)
        merged = merged.set_index("species_id").reindex(repaired["species_id"]).reset_index()
        out = repaired[INPUT_COLS].copy()
        for col in SURVEY_COLS:
            out[col] = merged[col].values
        for col in SURVEY_COLS:
            out[col] = out[col].map(lambda v: "" if is_null(v) else
                                    (str(int(float(v))) if str(v).replace(".0", "").lstrip("-").isdigit()
                                     else norm(v)))
        samples[taxon] = out

        print(f"  {taxon:<14} {len(src):>6} → {len(repaired):>6}행")

    # 두 시트가 모두 정비된 뒤에야 교차 참조를 맞출 수 있다
    fix_dual_habitat(masters, log)

    log_df = log.frame()
    write_workbook(OUT_MASTER, masters, _guide("마스터DB"), log_df)
    write_workbook(OUT_INPUT, inputs, _guide("입력용 데이터시트"))
    write_workbook(OUT_SAMPLE, samples, _guide("가상데이터"), log_df)

    print(f"\n정비 내역 {len(log_df)}건")
    if not log_df.empty:
        print(log_df["구분"].value_counts().to_string())
    for p in (OUT_MASTER, OUT_INPUT, OUT_SAMPLE):
        print(f"  {p.name}  ({p.stat().st_size / 1024 / 1024:.1f} MB)")


if __name__ == "__main__":
    main()
